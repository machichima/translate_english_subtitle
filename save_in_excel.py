import openpyxl as xl
import os

def save(vol, p_o_s, eng, ch, sentence, sentence_ch, len_p, path):
    name = open(os.getcwd()+'/data/store_place.txt')   #讀取store_place.txt中的資料
    place = path + '/' + name.read().split('\n')[3]   #讀取放入的excel的名稱
    p = os.path.exists(place)
    exist = 0
    if(p):
        wb = xl.load_workbook(place)   #讀取已有的excel
        ws = wb._sheets[0]
        exist = 1
    else: 
        wb = xl.Workbook()   #新增excel
        ws = wb._sheets[0]
        ws.title = 'vol'
    if(exist):   #excel已經存在，在已有資料的後面加入新的資料
        row = ws.max_row + 1
        ws.cell(row, 1).value = vol
        i = 0
        x = row
        while(i < len_p):
            ws.cell(x, 2).value = p_o_s[i]
            for j in range(x, x + len(eng[i])):
                y = j - x
                for k in range(0, 4):
                    ws.cell(j, k+3).value = [ch[i][y], eng[i][y], sentence[i][y], sentence_ch[i][y]][k]
            x = x + len(eng[i])
            i += 1
    else:   #新增excel，創建標頭，資料從第二行開始加
        for i in range(1, 6):
            ws.cell(1, i).value = ['volcabulary', '詞性', '中文解釋', '英文解釋', '例句'][i-1]
        ws.cell(2, 1).value = vol
        i = 0
        x = 2
        while(i < len_p):
            ws.cell(x, 2).value = p_o_s[i]
            for j in range(x, x + len(eng[i])):
                y = j - x
                for k in range(0, 4):
                    ws.cell(j, k+3).value = [ch[i][y], eng[i][y], sentence[i][y], sentence_ch[i][y]][k]
            x = x + len(eng[i])
            i += 1

    wb.save(place)
if(__name__ == '__main__'):
    vol = 'word'
    p_o_s = ['p_o_s']
    eng = ['eng']
    ch = ['中文']
    sentence = ['sentence'] 
    len_ch = 1
    quantity = 2
    save(vol, p_o_s, eng, ch, sentence, sentence_ch, len_eng, path)

